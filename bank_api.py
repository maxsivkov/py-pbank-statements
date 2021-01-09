from datadir import DataDir
from typing import List
import requests, re, datetime, dateutil
from itertools import chain
from functools import reduce
from operator import add

from dateutil.relativedelta import relativedelta
from dateutil.parser import parse
from decimal import Decimal

from openpyxl import Workbook

class ApiBase(object):
    QUARTER_FIRST_DATE = ['01-01', '04-01', '07-01', '10-01'] # MM.DD

    def __init__(self, d:DataDir):
        self.quarter_first_date=[]
        self._period_str=None
        self.period = self._period(d.get_config_v('statements.period.value'), d.get_config_v('statements.period.start'), d.get_config_v('statements.period.end'))
        self.ignore_accounts = d.get_config_v('statements.ignore_accounts')
        self.accounts:List[str] = []
        self._statements = {}

    def _period(self, value, start, end):
        first_date = None
        last_date = None

        if start and end:
            (st_year, st_month, st_day) = self._parse_date(start)
            (end_year, end_month, end_day) = self._parse_date(end)
            if st_year:
                first_date = datetime.date(st_year, st_month, st_day)
            if end_year:
                last_date = datetime.date(end_year, end_month, end_day)
            if not first_date:
                first_date = dateutil.parser.parse(start, fuzzy=True)
            if not last_date:
                last_date = dateutil.parser.parse(end, fuzzy=True)
            self._period_str = f'{first_date} -- {last_date}'
        elif value:
            if value.lower() == 'prev':
                first_date, last_date = self.prev_quarter(datetime.date.today())
                self._period_str=f'{first_date} -- {last_date}'
            else:
                (year, quarter, _) = self._parse_date(value)
                self._period_str = f'{year}-{quarter.lower()}'
                if not quarter.isdigit():
                    quarter = quarter.lower()
                    if quarter == 'i': quarter = 1
                    if quarter == 'ii': quarter = 2
                    if quarter == 'iii': quarter = 3
                    if quarter == 'iv': quarter = 4
                else: quarter = int(quarter)
                quarter_start = f'{year}-{ApiBase.QUARTER_FIRST_DATE[quarter-1]}'
                first_date = dateutil.parser.parse(quarter_start)
                last_date = first_date + relativedelta(months=3, days=-1)
        return first_date, last_date

    @property
    def first_date(self): return self.period[0]

    @property
    def last_date(self): return self.period[1]

    @property
    def period_str(self): return self._period_str

    def prev_quarter(self, ref):
        first_month_of_quarter = ((ref.month - 1) // 3) * 3 + 1
        last_date = ref.replace(month=first_month_of_quarter, day=1) - relativedelta(days=1)
        first_date = last_date - relativedelta(months=3, days=-1)
        return first_date, last_date

    def _parse_date(self, q):
        if not q: return q
        exps = [r'(?P<day>\d{1,2}).(?P<month>\d{1,2}).(?P<year>\d{4})',  #20.1.2020, 20-1-2020
                r'(?P<year>\d{4}).(?P<month>\d{1,2}).(?P<day>\d{1,2})',  # 2020.1.20, 2020-1-20
                r'(?P<year>\d{4}).(?P<quarter>\d)',  # 2020.4, 2020-4
                r'(?P<year>\d{4}).(?P<quarter>[iIvV]{1,2})',  # 2020.iv, 2020-iv
                r'(?P<quarter>\d).(?P<year>\d{4})',  # 4.2020, 4-2020
                r'(?P<quarter>[iIvV]{1,2}).(?P<year>\d{4})',  # iv.2020, iv-2020
                ]
        for r in exps:
            m = re.match(r, q)
            if m:
                values = m.groupdict()
                return (int(values['year']), values['quarter'], None) if 'quarter' in values else (int(values['year']), int(values['month']), int(values['day']) if "day" in values else None)
        return (None, None, None)

    def ignored(self, account):
        for ap in self.ignore_accounts:
            if re.match(ap, account) is not None:
                return True
        return False

class PbApi(ApiBase):
    def __init__(self, d:DataDir):
        ApiBase.__init__(self, d)

        self.pb_id:str = d.get_config_str('statements.pb_id')
        self.pb_token:str = d.get_config_str('statements.pb_token')

    def fetch_statements(self):
        self.accounts:List[str] = []

        params = {'startDate': self.first_date.strftime('%d-%m-%Y'), 'endDate': self.last_date.strftime('%d-%m-%Y')}
        headers = {'id': self.pb_id
            , 'token': self.pb_token
            , 'Content-type': 'application/json;charset=utf8'}
        r = requests.get("https://acp.privatbank.ua/api/proxy/transactions", params=params, headers=headers)
        r.raise_for_status()
        stmts = r.json()['StatementsResponse']['statements']
        self.accounts = list(chain.from_iterable(stmts))
        for account_item in stmts:
            self._statements.update(account_item.items())

    def balance(self):

        params = {'startDate': self.first_date.strftime('%d-%m-%Y'), 'endDate': self.last_date.strftime('%d-%m-%Y')}
        headers = {'id': self.pb_id
            , 'token': self.pb_token
            , 'Content-type': 'application/json;charset=utf8'}
        r = requests.get("https://acp.privatbank.ua/api/statements/balance/final", headers=headers)
        r.raise_for_status()
        return r.json()['balances']


    def statements(self, accountNo:str): return list(chain.from_iterable([x.values() for x in self._statements[accountNo]]))

    def datetime_pb(self, s) -> datetime : return datetime.datetime.strptime(s, '%d.%m.%Y %H:%M:%S')

    def store(self, wb_filename, start_row:int):
        all_statements = []
        for accountNo in self.accounts:
            if self.ignored(accountNo): continue
            all_statements.extend(self.statements(accountNo))
        all_statements.sort(key=lambda x: self.datetime_pb(x['DATE_TIME_DAT_OD_TIM_P']))
        wb = Workbook()
        ws = wb.active

        ws.cell(row=2, column=1, value=f'Виписка по декількох рахунках з {self.first_date} по {self.last_date}')

        #ws.title = f'{self.first_date}-{self.last_date}'
        ws.cell(row=4, column=1, value='№')
        ws.cell(row=4, column=2, value='Дата проводки')
        ws.cell(row=4, column=3, value='Час проводки')
        ws.cell(row=4, column=4, value='Сума')
        ws.cell(row=4, column=5, value='Валюта')
        ws.cell(row=4, column=6, value='Призначення платежу')
        ws.cell(row=4, column=7, value='ЄДРПОУ')
        ws.cell(row=4, column=8, value='Назва контрагента')
        ws.cell(row=4, column=9, value='Рахунок контрагента')
        ws.cell(row=4, column=10, value='МФО контрагента')
        ws.cell(row=4, column=11, value='Ваш МФО')
        ws.cell(row=4, column=12, value='Ваш ЄДРПОУ')
        ws.cell(row=4, column=13, value='Ваш рахунок')
        ws.cell(row=4, column=14, value='Назва вашого рахунку')
        ws.cell(row=4, column=15, value='Референс')


        row = start_row
        for s in all_statements:
            ws.cell(row=row, column=1, value=s['BPL_NUM_DOC'])
            date = self.datetime_pb(s['DATE_TIME_DAT_OD_TIM_P'])
            amount=Decimal(s['BPL_SUM'])
            ws.cell(row=row, column=2, value=date.strftime('%d.%m.%Y'))
            ws.cell(row=row, column=3, value=date.strftime('%H:%M:%S'))
            ws.cell(row=row, column=4, value=-amount if s['TRANTYPE'] == 'D' else amount)
            ws.cell(row=row, column=5, value=s['BPL_CCY'])
            ws.cell(row=row, column=6, value=s['BPL_OSND'])
            ws.cell(row=row, column=7, value=s['AUT_CNTR_CRF'])
            ws.cell(row=row, column=8, value=s['AUT_CNTR_NAM'])
            ws.cell(row=row, column=9, value=s['AUT_CNTR_ACC'])
            ws.cell(row=row, column=10, value=s['AUT_CNTR_MFO'])
            ws.cell(row=row, column=11, value=s['AUT_MY_MFO'])
            ws.cell(row=row, column=12, value=s['AUT_MY_CRF'])
            ws.cell(row=row, column=13, value=s['AUT_MY_ACC'])
            ws.cell(row=row, column=14, value=s['AUT_MY_NAM'])
            ws.cell(row=row, column=15, value=s['BPL_REF'])
            row=row+1
        wb.save(wb_filename)
        return len(all_statements)