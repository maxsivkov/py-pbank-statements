from typing import Dict
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.fills import PatternFill
from openpyxl.comments import Comment
from stmt_driver_base import StatementsDriverBase, StatementRow
from utils import *
from functools import partial

class XlsxDriver(StatementsDriverBase):
    TRANSACTION_ID_COL = 1

    def __init__(self, filepath:str, output_file_path:str=None, first_row:int = 0):
        super().__init__(filepath)
        self.first_row = first_row
        self.row_map:Dict[int, int] = {}
        self.filepath = filepath
        self.output_file_path = output_file_path
        self.filename = os.path.splitext(os.path.basename(self.filepath))[0]
        self.wb:Workbook = None
        self.ws:Worksheet = None
        self.max_rows:int = None

    def read(self):
        self.wb = load_workbook(self.filepath)
        self.wb.template = False
        self.ws = self.wb.worksheets[0]
        self.max_rows = self.ws.max_row
        for r in range(self.first_row, self.max_rows + 1):
            row = self.instantinate_row(r, lambda c : self.cell(r, c))
            row._drv_idx = r
            row.drvid = f'{self.filename}/{r}'
            row.result_ = partial(self.result, r)
                #lambda text : self.setter(r, 20, text)
            self.row_map[len(self.rows)] = r
            self.rows.append(row)

    def cell(self, r, c): return self.ws.cell(row=r, column=c).value

    def result(self, r, op_type, text):
        op_types = {'CurrencyExchange': 'FFADFF2F',
                    'FlowIncome': 'FF7CFC00',
                    'FlowOutgo': 'FF00FF00',
                    'Withdrawal': 'FF32CD32',
                    'unknown': 'FFE9967A'
                    }
        fill=PatternFill(start_color=op_types[op_type],
                   end_color=op_types[op_type],
                   fill_type='solid')
        cc = self.ws.cell(row=r, column=1)
        cc.comment = Comment(text, "py-bank-statements")
        #cc = self.ws['A1']
        #cc.fill = fill
        cc.fill = fill
        #cc.value = value
        #self.ws.cell(row=r, column=20, value=value)

    def close(self):
        #newfilepath = self.filename + '_out.xlsx'
        #newfilepath = os.path.join(os.path.dirname(os.path.abspath(self.filepath)), newfilepath)
        if not os.path.exists(os.path.dirname(self.output_file_path)):
            os.mkdir(os.path.dirname(self.output_file_path))

        if self.output_file_path is not None:
            self.wb.save(self.output_file_path)
        pass